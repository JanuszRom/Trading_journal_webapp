from flask import request, jsonify, current_app, send_from_directory
from werkzeug.utils import secure_filename
from models import db, Trade, Screenshot
import os
import re
import uuid
from datetime import datetime


def register_routes(app):
    # Get all trades
    @app.route('/api/trades', methods=['GET'])
    def get_trades():
        trades = Trade.query.order_by(Trade.timestamp.desc()).all()

        result = []
        for trade in trades:
            trade_data = {
                'id': trade.id,
                'timestamp': trade.timestamp.isoformat(),
                'instrument': trade.instrument,
                'direction': trade.direction,
                'entry': trade.entry,
                'exit': trade.exit,
                'stop_loss': trade.stop_loss,
                'take_profit': trade.take_profit,
                'size': trade.size,
                'risk': trade.risk,
                'reward': trade.reward,
                'profit_loss': trade.profit_loss,
                'duration': trade.duration,
                'comments': trade.comments,
                # 'strategy': trade.strategy,
                # 'setup': trade.setup,
                # 'mistakes': trade.mistakes,
                # 'lessons': trade.lessons,
                'screenshots': [{'id': s.id, 'filename': s.filename} for s in trade.screenshots]
            }
            result.append(trade_data)

        return jsonify(result)

    # Get a specific trade
    @app.route('/api/trades/<int:trade_id>', methods=['GET'])
    def get_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)

        trade_data = {
            'id': trade.id,
            'timestamp': trade.timestamp.isoformat(),
            'instrument': trade.instrument,
            'direction': trade.direction,
            'entry': trade.entry,
            'exit': trade.exit,
            'stop_loss': trade.stop_loss,
            'take_profit': trade.take_profit,
            'size': trade.size,
            'risk': trade.risk,
            'reward': trade.reward,
            'profit_loss': trade.profit_loss,
            'duration': trade.duration,
            'comments': trade.comments,
            # 'strategy': trade.strategy,
            # 'setup': trade.setup,
            # 'mistakes': trade.mistakes,
            # 'lessons': trade.lessons,
            'screenshots': [{'id': s.id, 'filename': s.filename} for s in trade.screenshots]
        }

        return jsonify(trade_data)

    # Create a new trade
    @app.route('/api/trades', methods=['POST'])
    def create_trade():
        data = request.form

        # Create new trade
        new_trade = Trade(
            instrument=data.get('instrument'),
            direction=data.get('direction'),
            entry=float(data.get('entry')),
            exit=float(data.get('exit')),
            stop_loss=float(data.get('stop_loss')),
            take_profit=float(data.get('take_profit')),
            size=float(data.get('size')),
            risk=float(data.get('risk')),
            reward=float(data.get('reward')),
            profit_loss=float(data.get('profit_loss')),
            duration=data.get('duration'),
            comments=data.get('comments', ''),
            # strategy=data.get('strategy'),
            # setup=data.get('setup', ''),
            # mistakes=data.get('mistakes', ''),
            # lessons=data.get('lessons', '')
        )

        db.session.add(new_trade)
        db.session.commit()

        # Handle screenshot uploads
        files = request.files.getlist('screenshots')
        for file in files:
            if file and file.filename:
                # Secure filename and make unique
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{filename}"

                # Save file
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(filepath)

                # Create screenshot record
                screenshot = Screenshot(
                    filename=unique_filename,
                    filepath=filepath,
                    trade_id=new_trade.id
                )
                db.session.add(screenshot)

        db.session.commit()

        return jsonify({'success': True, 'trade_id': new_trade.id}), 201

    # Update a trade
    @app.route('/api/trades/<int:trade_id>', methods=['PUT'])
    def update_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)
        data = request.form

        # Update trade fields
        trade.instrument = data.get('instrument', trade.instrument)
        trade.direction = data.get('direction', trade.direction)
        trade.entry = float(data.get('entry', trade.entry))
        trade.exit = float(data.get('exit', trade.exit))
        trade.stop_loss = float(data.get('stop_loss', trade.stop_loss))
        trade.take_profit = float(data.get('take_profit', trade.take_profit))
        trade.size = float(data.get('size', trade.size))
        trade.risk = float(data.get('risk', trade.risk))
        trade.reward = float(data.get('reward', trade.reward))
        trade.profit_loss = float(data.get('profit_loss', trade.profit_loss))
        trade.duration = data.get('duration', trade.duration)
        trade.comments = data.get('comments', trade.comments)
        # trade.strategy = data.get('strategy', trade.strategy)
        # trade.setup = data.get('setup', trade.setup)
        # trade.mistakes = data.get('mistakes', trade.mistakes)
        # trade.lessons = data.get('lessons', trade.lessons)

        # Handle new screenshots
        files = request.files.getlist('screenshots')
        for file in files:
            if file and file.filename:
                # Secure filename and make unique
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{filename}"

                # Save file
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(filepath)

                # Create screenshot record
                screenshot = Screenshot(
                    filename=unique_filename,
                    filepath=filepath,
                    trade_id=trade.id
                )
                db.session.add(screenshot)

        db.session.commit()

        return jsonify({'success': True})

    # Delete a trade
    @app.route('/api/trades/<int:trade_id>', methods=['DELETE'])
    def delete_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)

        # Delete associated screenshots
        for screenshot in trade.screenshots:
            # Delete file from disk
            try:
                os.remove(screenshot.filepath)
            except:
                pass

            # Delete record
            db.session.delete(screenshot)

        # Delete trade
        db.session.delete(trade)
        db.session.commit()

        return jsonify({'success': True})

    # Get screenshot
    @app.route('/api/screenshots/<filename>')
    def get_screenshot(filename):
        return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

    # Delete screenshot
    @app.route('/api/screenshots/<int:screenshot_id>', methods=['DELETE'])
    def delete_screenshot(screenshot_id):
        screenshot = Screenshot.query.get_or_404(screenshot_id)

        # Delete file from disk
        try:
            os.remove(screenshot.filepath)
        except:
            pass

        # Delete record
        db.session.delete(screenshot)
        db.session.commit()

        return jsonify({'success': True})

    # Export to Excel endpoint
    @app.route('/api/export/excel')
    def export_excel():
        import io
        import os
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file

        file_path = os.path.join(app.root_path, 'static', 'exports', 'trades_export.xlsx')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Check if file exists
        if os.path.exists(file_path):
            # Load existing workbook
            wb = load_workbook(file_path)
            ws = wb.active if "Trade Journal" in wb.sheetnames else wb.create_sheet("Trade Journal")
            
            # Get existing trade IDs from the Excel file
            existing_ids = set()
            id_col = None
            
            # Find the ID column if it exists
            for idx, cell in enumerate(ws[1]):
                if cell.value == 'ID':
                    id_col = idx
                    break
            
            # If ID column exists, collect existing IDs
            if id_col is not None:
                for row in ws.iter_rows(min_row=2):
                    if row[id_col].value:
                        existing_ids.add(row[id_col].value)
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trade Journal"
            
            # Add headers
            headers = [
                'ID', 'Timestamp', 'Instrument', 'Direction', 'Entry', 'Exit',
                'Stop Loss', 'Take Profit', 'Size', 'Risk', 'Reward',
                'P/L', 'Duration', 'comments',
                # 'Strategy', 'Setup', 'Mistakes', 'Lessons'
            ]
            ws.append(headers)

            # Format headers
            header_style = {
                'fill': PatternFill(start_color='4F81BD', fill_type='solid'),
                'font': Font(color='FFFFFF', bold=True),
                'alignment': Alignment(horizontal='center')
            }

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                for attr, value in header_style.items():
                    setattr(cell, attr, value)
                    
            existing_ids = set()

        # Add data
        trades = Trade.query.all()
        next_row = ws.max_row + 1
        
        for trade in trades:
            # Skip if trade already exists in the file
            if hasattr(trade, 'id') and trade.id in existing_ids:
                continue
                
            row = [
                trade.id if hasattr(trade, 'id') else None,
                trade.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                trade.instrument,
                trade.direction,
                trade.entry,
                trade.exit,
                trade.stop_loss,
                trade.take_profit,
                trade.size,
                trade.risk,
                trade.reward,
                trade.profit_loss,
                trade.duration,
                trade.comments
                # trade.strategy,
                # trade.setup,
                # trade.mistakes,
                # trade.lessons
            ]
            ws.append(row)

        # Adjust column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Save to disk
        wb.save(file_path)
        
        # Also save to memory for download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='trades_export.xlsx',
            as_attachment=True
        )

    

    @app.route('/api/parse-trade', methods=['POST'])
    def parse_trade():
        try:
            data = request.get_json(silent=True) or {}
            clipboard_text = data.get('clipboard_text', '')
            if not clipboard_text:
                return jsonify({'error': 'No clipboard_text provided'}), 400
            
            try:
                trade_data = parse_xstation5_format(clipboard_text)
                print(f"Parsed trade data: {trade_data}")  # Add debug log
                return jsonify({'trade': trade_data})
            except Exception as e:
                print(f"Error parsing trade: {str(e)}")  # Add debug log
                return jsonify({'error': str(e)}), 400
        except Exception as e:
            print("Error parsing trade:", e)
            return jsonify({"error": str(e)}), 500
    
    def parse_xstation5_format(text):
        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
        
        trade = {}
        label_map = {
            "Typ": "direction",
            "Wolumen": "size",
            "Zysk netto": "profit_loss",
            "Cena otwarcia": "entry",
            "Cena zamknięcia": "exit",
            "Stop Loss": "stop_loss",
            "Take Profit": "take_profit",
        }

        # Detect instrument (first non-empty line after "Szczegóły pozycji")
        try:
            idx = lines.index("Szczegóły pozycji")
            instrument = lines[idx + 1]
            trade["instrument"] = instrument
        except Exception:
            trade["instrument"] = ""
        def parse_float(value):
            """Convert localized float like '18 939.71' to float 18939.71"""
            try:
                return float(value.replace(" ", "").replace(",", "."))
            except:
                return 0.0
        # Walk through all lines to get data
        i = 0
        while i < len(lines):
            line = lines[i]
            if line in label_map:
                key = label_map[line]
                value = lines[i + 1] if i + 1 < len(lines) else ""
                if key == "direction":
                    trade["direction"] = "long" if value.lower() == "buy" else "short"
                elif key in ["entry", "exit", "stop_loss", "take_profit", "profit_loss", "size"]:
                    trade[key] = parse_float(value)
                i += 2
            elif line == "Czas otwarcia":
                if i + 2 < len(lines):
                    open_time = f"{lines[i + 1]} {lines[i + 2]}"
                    trade["open_time"] = open_time
                    i += 3
                else:
                    i += 1
            elif line == "Czas zamknięcia":
                if i + 2 < len(lines):
                    close_time = f"{lines[i + 1]} {lines[i + 2]}"
                    trade["close_time"] = close_time
                    i += 3
                else:
                    i += 1
            else:
                i += 1

        # Calculate duration
        try:
            if "open_time" in trade and "close_time" in trade:
                fmt = "%d.%m.%Y %H:%M"
                t1 = datetime.strptime(trade["open_time"], fmt)
                t2 = datetime.strptime(trade["close_time"], fmt)
                delta = t2 - t1
                mins = int(delta.total_seconds() // 60)
                trade["duration"] = f"{mins} min"
        except Exception:
            trade["duration"] = ""

        try:
            entry = trade.get("entry", 0)
            sl = trade.get("stop_loss", 0)
            tp = trade.get("take_profit", 0)
            size = trade.get("size", 0)
            direction = trade.get("direction", "")
            instrument = trade.get("instrument", "")
            
            # Check if instrument is forex (common forex pairs contain currency codes)
            is_forex = any(pair in instrument.upper() for pair in ["USD", "EUR", "GBP", "JPY", "AUD", "NZD", "CAD", "CHF"])
    
            if all([entry, sl, tp, size]):
                if direction == "long":
                    if is_forex:
                        # Forex calculation using pips
                        risk_pips = abs(entry - sl) * 10000  # Convert to pips
                        reward_pips = abs(tp - entry) * 10000
                        trade["risk"] = round(risk_pips * size, 2)
                        trade["reward"] = round(reward_pips * size, 2)
                    else:
                        # Original calculation for other instruments
                        trade["risk"] = round((entry - sl) * size, 2)
                        trade["reward"] = round((tp - entry) * size, 2)
                elif direction == "short":
                    if is_forex:
                        # Forex calculation using pips
                        risk_pips = abs(sl - entry) * 10000
                        reward_pips = abs(entry - tp) * 10000
                        trade["risk"] = round(risk_pips * size, 2)
                        trade["reward"] = round(reward_pips * size, 2)
                    else:
                        # Original calculation for other instruments
                        trade["risk"] = round((sl - entry) * size, 2)
                        trade["reward"] = round((entry - tp) * size, 2)
        except Exception as e:
            print(f"Error calculating risk/reward: {e}")
            trade["risk"] = 0
            trade["reward"] = 0

        return trade
   # Add this endpoint to your existing routes.py file

    # Upload screenshot for existing trade
    @app.route('/api/trades/<int:trade_id>/screenshots', methods=['POST'])
    def upload_screenshot(trade_id):
        trade = Trade.query.get_or_404(trade_id)
        
        if 'screenshot' not in request.files:
            return jsonify({'error': 'No screenshot file provided'}), 400
        
        file = request.files['screenshot']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and file.filename:
            # Secure filename and make unique
            filename = secure_filename(file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"

            # Save file
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(filepath)

            # Create screenshot record
            screenshot = Screenshot(
                filename=unique_filename,
                filepath=filepath,
                trade_id=trade.id
            )
            db.session.add(screenshot)
            db.session.commit()
            
            # Return screenshot data for frontend
            return jsonify({
                'id': screenshot.id,
                'filename': screenshot.filename,
                'filepath': screenshot.filepath,
                'upload_date': screenshot.upload_date.isoformat(),
                'trade_id': screenshot.trade_id
            }), 201
        
        return jsonify({'error': 'Invalid file'}), 400